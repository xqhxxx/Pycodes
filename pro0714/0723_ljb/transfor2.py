# 打开文件：
import openpyxl
from openpyxl import load_workbook

# 打开读取的表
excel = load_workbook('D:/Pycodes/pro0714/tools/0723_ljb/aaa.xlsx')
# 获取sheet：
table = excel.get_sheet_by_name('词汇表')  # 通过表名获取
# 获取行数和列数：
rows = table.max_row  # 获取行数
cols = table.max_column  # 获取列数
# 获取单元格值：
# Data = table.cell(row=2, column=1).value  # 获取表格内容，是从第一行第一列是从1开始的，注意不要丢掉 .value

# 创建 list放词汇映射的各个等级
list_codes = []
# 创建字典 存放每个等级的 k-v映射
dict_c_1 = {}
dict_c_2 = {}
dict_c_3 = {}
dict_c_4 = {}
dict_c_5 = {}
dict_c_6 = {}

# 遍历每一行
for i in range(2, rows):
    # 获取每一行的映射等级
    data_type = table.cell(row=i, column=4).value
    # 根据等级放在相对应的字典中
    if '1' == str(data_type):
        dict_c_1[str(table.cell(row=i, column=2).value)] = table.cell(row=i, column=3).value
    if '2' == str(data_type):
        dict_c_2[str(table.cell(row=i, column=2).value)] = table.cell(row=i, column=3).value
    if '3' == str(data_type):
        dict_c_3[str(table.cell(row=i, column=2).value)] = table.cell(row=i, column=3).value
    if '4' == str(data_type):
        dict_c_4[str(table.cell(row=i, column=2).value)] = table.cell(row=i, column=3).value
    if '5' == str(data_type):
        dict_c_5[str(table.cell(row=i, column=2).value)] = table.cell(row=i, column=3).value
    if '6' == str(data_type):
        dict_c_6[str(table.cell(row=i, column=2).value)] = table.cell(row=i, column=3).value
# 将字典放进list
list_codes.append(dict_c_1)
list_codes.append(dict_c_2)
list_codes.append(dict_c_3)
list_codes.append(dict_c_4)
list_codes.append(dict_c_5)
list_codes.append(dict_c_6)

# 读取原数据表 并替换成eng
# 获取sheet：
table_2 = excel.get_sheet_by_name('数据对象表')  # 通过表名获取
# 获取行数：
rows_2 = table_2.max_row  # 获取行数

'''
新建一个表来存放转换的数据
'''
wb = openpyxl.Workbook()
# 当前打开的sheet页 wb.active
ws = wb.active
# #更改默认名称Sheet`
ws.title = "数据对象表"

print("开始转换")
# 遍历源数据表的每一行
for i in range(2, rows_2):
    chi_value = table_2.cell(row=i, column=1).value
    eng_vaule = table_2.cell(row=i, column=1).value
    # 遍历映射 按等级
    for codes in list_codes:
        for key in codes:#遍历映射表
            if eng_vaule is not None:
                if key in eng_vaule:#判断原数据含有映射字段
                    if codes[key] is not None:
                        # 有就替换为eng
                        eng_vaule = eng_vaule.replace(key, codes[key])
    #替换完成的数据放进新表的单元格中
    ws.cell(i, 1).value = chi_value
    ws.cell(i, 2).value = eng_vaule
    print(i, " 行", eng_vaule)
    # if (i%5000==0 or i==rows_2):

wb.save('D:/Pycodes/pro0714/tools/0723_ljb/bbb.xlsx')


